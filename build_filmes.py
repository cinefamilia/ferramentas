#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_filmes.py — Cine Família
Reconstruído do zero na sessão B09.2, seguindo as regras acumuladas
documentadas nas Instruções do Projeto / memória de sessões anteriores.

Lê Cine_Fami_lia_B09_2.xlsx e gera filmes.js (window.FILMES = [...]).
"""

import json
import openpyxl

XLSX_PATH = "Cine_Fami_lia_B09_2.xlsx"
OUT_PATH = "filmes.js"

# Abas que são "gênero" puro -> genero = nome da aba, origem = "genero"
# Coleções e Cine Itália têm tratamento próprio.
ABA_COLECOES = "Coleções"
ABA_CINE_ITALIA = "Cine Itália"

# Mapa fixo para os 4 filmes de dupla coleção (gênero + Belle Cult) cujo
# campo Nome NÃO segue o padrão "Artista - Título" (o nome é o título puro).
ARTISTA_HARDCODE = {
    "Aeroporto 79 - O Concorde": "Sylvia Kristel",
    "Django": "Franco Nero",
    "O Dólar Furado": "Giuliano Gemma",
    "A Noite Americana": "Jacqueline Bisset",
}


def s(v):
    """strip seguro (None -> None, tudo mais -> str stripada ou None se vazia)."""
    if v is None:
        return None
    v = str(v).strip()
    return v if v != "" else None


def parse_tags(colecao_tipo):
    """Divide o campo 'Coleção Tipo' (pode ter múltiplas tags separadas por ';')
    em uma lista de objetos {tipo, nome, raw}."""
    if not colecao_tipo:
        return []
    partes = [p.strip() for p in colecao_tipo.split(";") if p.strip()]
    tags = []
    for parte in partes:
        if parte.startswith("Franquia - "):
            tags.append({
                "tipo": "Franquia",
                "nome": parte[len("Franquia - "):].strip(),
                "raw": parte,
            })
        elif parte.startswith("Coleção "):
            tags.append({
                "tipo": "Coleção",
                "nome": parte,
                "raw": parte,
            })
        else:
            tags.append({
                "tipo": "Outro",
                "nome": parte,
                "raw": parte,
            })
    return tags


def is_belle_cult(tags):
    return any("Belle Cult" in t["raw"] for t in tags)


def get_hyperlink_or_value(cell):
    if cell.hyperlink is not None and cell.hyperlink.target:
        return s(cell.hyperlink.target)
    return s(cell.value)


def build_row(headers, row_cells, aba_nome, origem, genero):
    idx = {h: i for i, h in enumerate(headers)}

    def val(col):
        return row_cells[idx[col]]

    nome_raw = val("Nome").value
    if nome_raw is None:
        return None  # linha vazia
    nome = s(str(nome_raw))  # força string (títulos numéricos tipo "300")

    capa = get_hyperlink_or_value(val("Capa"))
    nome_original = s(val("Nome Original").value)
    ano = val("Ano").value
    duracao = s(val("Duração").value)  # NUNCA normalizar
    gb = val("GB").value
    local = s(val("Local").value)
    codec = s(val("Codec").value)
    perfil_cor = s(val("Perfil Cor").value)
    preset_elmedia = s(val("Preset Elmedia").value)
    situacao = s(val("Situação").value)
    sinopse = s(val("Sinopse").value)
    colecao_tipo = s(val("Coleção Tipo").value)
    ordem = val("Ordem").value
    if isinstance(ordem, str):
        ordem = s(ordem)
    diretor = s(val("Diretor").value)
    atores = [s(val(f"Ator{i}").value) for i in range(1, 7)]

    tags = parse_tags(colecao_tipo)

    artista = None
    if is_belle_cult(tags):
        if nome in ARTISTA_HARDCODE:
            artista = ARTISTA_HARDCODE[nome]
            # nome permanece como está (não é padrão "Artista - Título")
        elif " - " in nome:
            esquerda, direita = nome.rsplit(" - ", 1)
            artista = esquerda.strip()
            nome = direita.strip()

    row = {
        "capa": capa,
        "nome": nome,
        "ano": ano,
        "duracao": duracao,
        "gb": gb,
        "local": local,
        "codec": codec,
        "perfilCor": perfil_cor,
        "presetElmedia": preset_elmedia,
        "situacao": situacao,
        "sinopse": sinopse,
        "genero": genero,
        "origem": origem,
        "colecaoTipo": colecao_tipo,
        "ordem": ordem,
        "tags": tags,
        "diretor": diretor,
        "ator1": atores[0],
        "ator2": atores[1],
        "ator3": atores[2],
        "ator4": atores[3],
        "ator5": atores[4],
        "ator6": atores[5],
        "nomeOriginal": nome_original,
    }
    if artista:
        row["artista"] = artista
    return row


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    filmes = []

    for aba_nome in wb.sheetnames:
        ws = wb[aba_nome]
        headers = [c.value for c in ws[1]]

        if aba_nome == ABA_COLECOES:
            origem, genero = "colecoes", None
        elif aba_nome == ABA_CINE_ITALIA:
            origem, genero = "cine_italia", None
        else:
            origem, genero = "genero", aba_nome

        for r in range(2, ws.max_row + 1):
            row_cells = ws[r]
            row = build_row(headers, row_cells, aba_nome, origem, genero)
            if row is not None:
                filmes.append(row)

    with open(OUT_PATH, "w", encoding="utf-8") as f:
        f.write("// Gerado automaticamente por build_filmes.py — não editar manualmente.\n")
        f.write("const filmes = ")
        f.write(json.dumps(filmes, ensure_ascii=False, indent=2))
        f.write(";\n\n")
        f.write("if (typeof module !== 'undefined') { module.exports = filmes; }\n")

    print(f"Gerados {len(filmes)} filmes em {OUT_PATH}")


if __name__ == "__main__":
    main()
